#!/usr/bin/env python
# coding: utf-8

"""
Refactoring and Documentation Standards
Documentation:

The documentation uses the NumPy style, which includes a clear breakdown of parameters, return values, and notes. This style is chosen for its structured format, which is highly readable and often preferred in scientific and machine learning codebases.
Each parameter is thoroughly explained, including types and example values.
The "Notes" section provides additional context on the function’s behavior and specific configurations.
Code Refactoring:

PEP8 Compliance: The refactoring follows PEP8 standards, enhancing readability by avoiding long lines, ensuring consistent indentation, and spacing around operators and comments.
Modularization and Clarity:
Variables are descriptively named (conv_output, optimizer_name, learning_rate), enhancing code clarity.
The conditional structure for selecting the optimizer is refactored for readability, and an error is raised for unsupported optimizers, adding robustness.
Comments in Portuguese: Comments clarify code steps in Portuguese, as requested, supporting native language readability while keeping documentation in English for broader accessibility.
This revised function adheres to best practices for readability, maintainability, and international usability.

"""

# # import

# In[1]:

import keras

from keras.models import Model
from keras import layers
from keras.applications import ResNet50, MobileNet, DenseNet201, InceptionV3
from keras.applications import ResNet152V2, Xception
import pandas as pd
import openpyxl
import sys,os
import argparse

import tensorflow as tf

os.environ["tf_gpu_allocator"]="cuda_malloc_async"
# Add the current directory to the PYTHONPATH
sys.path.insert(0, os.getcwd())
print(sys.path)

from models import  utils, get_data, models_train, sound_test_finalizado, maneger_gpu, send_whatsApp_msn
from models import reports_build as reports


# In[2]:
    
def print_layer(conv_model, layers_params):
    """
    Prints and saves information about the layers of a convolutional model.

    Parameters
    ----------
    conv_model : keras.Model
        The convolutional model whose layer details will be printed and optionally saved.
    layers_params : dict
        Dictionary containing parameters for layer information storage, including:
            - 'save_dir' : str
                Directory path where the layer information will be saved.
            - 'id_test' : str
                Unique identifier for the test or model instance.
            - 'model' : str
                Model name to include in the saved file.

    Returns
    -------
    None
        This function prints layer details and, if a directory is specified, saves it to a CSV file.
    
    Notes
    -----
    If 'save_dir' is provided, this function creates a directory structure in 'save_dir/models/' 
    and saves the CSV file named '<id_test>_<model>_layers.csv'. The file includes the 
    trainable status and name of each layer in the model.
    """
    # Define save directory based on parameters
    save_dir = layers_params['save_dir']
    # Create a model name using test ID and model name from parameters
    nm_model = f"{layers_params['id_test']}_{layers_params['model']}"
    
    if save_dir:
        # Specify the full path to the model save directory
        save_dir = f"{save_dir}/{nm_model}/"
        print('save_dir ', save_dir)
        # Create necessary folders in the specified save directory
        utils.create_folders(save_dir, flag=0)
        
        # Define CSV path for saving layer information
        _csv_layers = save_dir + '/' + nm_model + '_layers.csv'
        
        # Add initial test ID and model name to the CSV
        utils.add_row_csv(_csv_layers, [['id_test', layers_params['id_test']]])
        utils.add_row_csv(_csv_layers, [['model', layers_params['model']]])
        utils.add_row_csv(_csv_layers, [['trainable', 'name']])

    # Initialize an array to store layer details
    layers_arr = []
    # Iterate through each layer in the model
    for i, layer in enumerate(conv_model.layers):
        # Print the layer index, trainable status, and name
        print("{0} {1}:\t{2}".format(i, layer.trainable, layer.name))
        layers_arr.append([layer.trainable, layer.name])
    
    # Save layer details to CSV if save directory was specified
    if save_dir:
        utils.add_row_csv(_csv_layers, layers_arr)
    

# In[3]:

# Channel Attention Module    
def Channel(input_shape, r):
    # Calculate average and maximum values along height and width dimensions
    avg_p = tf.reduce_mean(input_shape, axis=[1, 2], keepdims=True)
    max_p = tf.reduce_max(input_shape, axis=[1, 2], keepdims=True)
    
    reduced_features = int(input_shape.shape[-1] // r) # Calculate the reduced number of features
    
    # Define dense layers
    dense1 = layers.Dense(reduced_features, activation="relu")
    dense2 = layers.Dense(input_shape.shape[-1], activation="linear")
    
    # Pass average and maximum values through the first dense layer
    Dense1_avg = dense1(avg_p)
    Dense1_max = dense1(max_p)

    # Pass the outputs of the first dense layer through the second dense layer
    Dense2_avg = dense2(Dense1_avg)
    Dense2_max = dense2(Dense1_max)

    # Apply sigmoid activation to the sum of outputs of the second dense layer
    out = layers.Activation('sigmoid')(Dense2_avg + Dense2_max)

    return out * input_shape # Scale the input_shape by the sigmoid output

# Spatial Attention Module
def Spatial(input_shape, ks=7):
    # Calculate average and maximum values along the channel dimension
    avg_p = tf.reduce_mean(input_shape, axis=[-1], keepdims=True)
    max_p = tf.reduce_max(input_shape, axis=[-1], keepdims=True)

    # Concatenate average and maximum values along the channel dimension
    concat_pool = tf.concat([avg_p, max_p], axis=-1)
    
    # Apply convolutional layer with kernel size (ks, ks) and sigmoid activation
    out = layers.Conv2D(1, (ks, ks), padding='same', activation='sigmoid')(concat_pool)

    # Scale the input_shape by the sigmoid output
    return out * input_shape

# CBAM
def CBAM(input_shape, ks=4, r=2):
    # Apply channel attention mechanism
    channel_out = Channel(input_shape, r)
    
    # Apply spatial attention mechanism on the output of channel attention
    Spatial_out = Spatial(channel_out, ks)
    
    # Add the output of channel attention and spatial attention
    out = layers.Add()([channel_out, Spatial_out])
    
    return layers.Activation('relu')(out) # Apply ReLU activation to the combined output


# ## hyper_model
def hyper_model(config_model):
    """
    Builds and configures a fine-tuned model based on a pre-trained base model.

    Parameters
    ----------
    config_model : dict
        Configuration dictionary with the following keys:
            - 'model' : str
                Name of the pre-trained model to be used (e.g., "VGG16").
            - 'id_test' : str
                Identifier for the test or specific model instance.
            - 'num_classes' : int
                Number of output classes for the classification task.
            - 'last_activation' : str
                Activation function for the final dense layer (e.g., "softmax").
            - 'freeze' : int
                Number of layers to freeze in the base model for transfer learning.
            - 'save_dir' : str
                Path to the directory where layer information will be saved.
            - 'optimizer' : str
                Name of the optimizer to use (e.g., "Adam", "SGD", "RMSprop", "Adagrad").
            - 'learning_rate' : float
                Learning rate for the optimizer.

    Returns
    -------
    keras.Model
        Compiled Keras model with fine-tuning and the specified configuration.

    Notes
    -----
    This function loads a pre-trained model (with 'imagenet' weights), freezes a certain number
    of layers as specified, adds a custom dense layer for classification, and optionally unfreezes
    some layers for further fine-tuning. The optimizer and learning rate are also set according 
    to the provided configuration.

    Documentation Style
    -------------------
    - Function documentation follows the **NumPy style** for readability and structured presentation.
    - Code is refactored according to **PEP8** coding standards for Python, focusing on readability,
      modularity, and clear comments.
    """
    
    # Initialize the specified pre-trained model
    #model = eval(config_model['model'])
    id_test = config_model['id_test']
    input_shape = config_model['input_shape']

    MainInput=layers.Input(shape=input_shape)
    
    #block-1 64
    fine_model=(layers.Conv2D(filters=64,kernel_size=(3,3), activation="relu",
                       padding="same",name="block1_conv1")(MainInput))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=64,kernel_size=(3,3), activation="relu",
                       padding="same",name="block1_conv2")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    
    fine_model=(layers.MaxPooling2D(strides=(2, 2),padding="same", 
                             name="block1_Max_polling2D")(fine_model))

    #block-2 128
    fine_model=(layers.Conv2D(filters=128,kernel_size=(3,3), activation="relu",
                       padding="same",name="block2_conv1")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=128,kernel_size=(3,3), activation="relu",
                       padding="same",name="block2_conv2")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=256,kernel_size=(3,3), activation="relu",
                       padding="same",name="block2_conv3")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    
    fine_model=(layers.MaxPooling2D(strides=(2, 2),padding="same",
                             name="block2_Max_polling2D")(fine_model))

    #block-3 256
    fine_model=(layers.Conv2D(filters=256,kernel_size=(3,3), activation="relu",
                       padding="same",name="block3_conv1")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=256,kernel_size=(3,3), activation="relu",
                       padding="same",name="block3_conv2")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=256,kernel_size=(3,3), activation="relu",
                       padding="same",name="block3_conv3")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=256,kernel_size=(3,3), activation="relu",
                       padding="same",name="block3_conv4")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    
    fine_model=(layers.MaxPooling2D(strides=(2, 2),padding="same",
                             name="block3_Max_polling2D")(fine_model))
    
    #block-4 512
    fine_model=(layers.Conv2D(filters=512,kernel_size=(3,3), activation="relu",
                       padding="same",name="block4_conv1")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=512,kernel_size=(3,3), activation="relu",
                       padding="same",name="block4_conv2")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=512,kernel_size=(3,3), activation="relu",
                       padding="same",name="block4_conv3")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))
    fine_model=(layers.Conv2D(filters=512,kernel_size=(3,3), activation="relu",
                       padding="same",name="block4_conv4")(fine_model))
    fine_model=(layers.BatchNormalization()(fine_model))

    
    fine_model = CBAM(fine_model, ks=config_model['ks'], r=config_model['r'])  # Apply CBAM
    
    fine_model = layers.GlobalAveragePooling2D()(fine_model) # Global average pooling layer
    
    # Fully connected layer with num_classes nodes and softmax activation
    #fine_model = Dense(num_classes, activation='softmax')(fine_model)
    
    
    fine_model=layers.Dense(config_model['num_classes'], 
                            activation=config_model['last_activation'])(fine_model)
    
    model = Model(inputs=MainInput, outputs=fine_model)

    # Prepare parameters for saving layer information
    layers_params = {'id_test': id_test, 'save_dir': config_model['save_dir'], 'model': config_model['model']}
    
    # Print and save layer details
    print_layer(model, layers_params)

    # Set the optimizer based on configuration
    optimizer_name = config_model['optimizer']
    learning_rate = config_model['learning_rate']
    if optimizer_name == 'Adam':
        opt = keras.optimizers.Adam(learning_rate=learning_rate)
    elif optimizer_name == 'RMSprop':
        opt = keras.optimizers.RMSprop(learning_rate=learning_rate)
    elif optimizer_name == 'Adagrad':
        opt = keras.optimizers.Adagrad(learning_rate=learning_rate)
    elif optimizer_name == 'SGD':
        opt = keras.optimizers.SGD(learning_rate=learning_rate)
    else:
        raise ValueError(f"Unsupported optimizer: {optimizer_name}")
    print(opt)

    # Compile the model with categorical crossentropy loss and accuracy metric
    model.compile(loss="categorical_crossentropy", optimizer=opt, metrics=["accuracy"])
    
    model.summary()
    return model


# In[4]:
    
def sheet_exists(path_xlsx):
    """
    Checks if a specified sheet exists in an Excel workbook. If the sheet is not present, it creates it 
    and initializes the first row with column headers.

    Parameters
    ----------
    path_xlsx : str
        The file path of the Excel workbook to check or create the sheet in.

    Returns
    -------
    None
        The function prints the names of sheets in the workbook and creates the sheet with specified 
        columns if it does not already exist.

    Notes
    -----
    This function uses openpyxl to handle Excel files. It checks for the sheet named 'Table_Exp'. If
    found, it accesses the sheet. If not found, it creates the sheet and adds headers. The workbook 
    is then saved to persist any changes.

    Raises
    ------
    FileNotFoundError
        If the specified Excel file does not exist, a message is printed.
    Exception
        Any other error during the file operation is captured and printed.
    """
    
    try:
        # Load the Excel workbook
        rel_book = openpyxl.load_workbook(path_xlsx)
        print(rel_book.sheetnames)  # Visualiza as abas existentes
        
        # Define columns for the sheet header
        cols_exe = [
            'id_test', 'nm_model', 'k', 'val_loss', 'val_acc', 
            'test_loss', 'test_accuracy', 'precision', 'recall', 
            'fscore', 'kappa', 'start_time', 'end_time', 
            'duration', 'best_epoch', 'num_eapoch', 'algorithm'
        ]
        
        _teste = 'Table_Exp'
        print('\n_test ', _teste)
        
        # Check if the sheet already exists
        if _teste in rel_book.sheetnames:
            print('Sheet exists.')
            Met_page = rel_book[_teste]  # Access the existing sheet
        else:
            print('Creating new sheet.')
            Met_page = rel_book.create_sheet(_teste)  # Create a new sheet
            Met_page.append(cols_exe)  # Add header row with column names
            
        # Save the workbook after adding or accessing the sheet
        rel_book.save(path_xlsx)
        print(rel_book.sheetnames)  # Visualiza as abas existentes
        
    except FileNotFoundError:
        print(f"Error: File '{path_xlsx}' not found.")
    except Exception as e:
        print(f"An error occurred while reading the file: {e}")


def read_sheet(path_xlsx):
    """
    Reads the "Params" and "Test" sheets from an Excel file and displays their content.

    Parameters
    ----------
    path_xlsx : str
        Path to the Excel file (.xlsx) containing the "Params" and "Test" sheets.

    Returns
    -------
    tuple
        A tuple containing:
        - config : pd.Series
            A series with the configuration parameters from the "Params" sheet.
        - rows : pd.DataFrame
            A DataFrame with test data from the "Test" sheet.
        If an error occurs, returns None.

    Raises
    ------
    FileNotFoundError
        If the specified file is not found, this error is raised.
    Exception
        Raised for other file reading or data processing errors.

    Functionality
    -------------
    - Reads the "Params" sheet, transposes it, and displays its parameters and values.
    - Reads the "Test" sheet and displays the test data stored in it.

    Examples
    --------
    >>> config, rows = read_sheet("path/to/file.xlsx")
    >>> print(config["epochs"])
    """
    try:
        # Carregar e transpor a aba "Params" com a primeira coluna como índice
        df = pd.read_excel(path_xlsx, sheet_name="Params", header=None)
        dft = df.set_index(0).T  # Define a primeira coluna como índice e transpõe
        config = dft.iloc[0]  # Extrair a primeira linha como série de configuração

        # Verificar se o DataFrame de configuração não está vazio
        if not config.empty:
            print("\nparams - values (first column as header)")
            # Exibir o conteúdo da aba "Params" transposta
            print(config)
            print("Epocas================", config["epochs"])

        # Carregar a aba "Test" do arquivo Excel
        rows = pd.read_excel(path_xlsx, sheet_name="Test")

        # Verificar se o DataFrame de teste não está vazio
        if not rows.empty:
            # Exibir os dados de teste armazenados
            print("\nDisplaying the registered tests")
            for i, row in rows.iterrows():
                print(row)

        return config, rows

    except FileNotFoundError:
        print(f"Error: File '{path_xlsx}' not found.")
    except Exception as e:
        print(f"An error occurred while reading the file: {e}")

        
def save_sheet(path_xlsx, data, sheet_name):
    """
    Save data to a specific sheet in an Excel file.

    Parameters
    ----------
    path_xlsx : str
        The path to the Excel file.
    data : list or tuple
        The data to be appended to the sheet.
    sheet_name : str
        The name of the sheet where the data will be inserted.

    Returns
    -------
    None
        The function does not return any value. It saves the data to the specified sheet in the Excel file.

    Notes
    -----
    The function assumes that the Excel file and the specified sheet already exist.
    """
    rel_book = openpyxl.load_workbook(path_xlsx)
    Met_page = rel_book[sheet_name]  # inserir dados
    Met_page.append(data)  # primeira linha
    rel_book.save(path_xlsx)

    


# In[6]:

def process_categories(row, k, verbose=0):
    """
    Process and retrieve the list of categories for a given k-fold.

    Parameters
    ----------
    row : pandas.Series
        A row from a DataFrame containing data for each test execution.
    k : int
        The k-fold number.
    verbose : int, optional
        The verbosity level (default is 1). If verbose > 0, the function prints information.

    Returns
    -------
    list
        A sorted list of category names.

    Notes
    -----
    The function prints information about the k-fold number, path data, categories,
    and the number of classes if verbose > 0. It retrieves category names from the
    specified path and returns a sorted list of these categories.
    """
    if verbose > 0:
        print("\n1-[INFO] k-fold number ", k)
    
    _path_data = row['path_data']
    
    if verbose > 0:
        print('path_data ', _path_data)
    
    _path_cat = f"{_path_data}/Train/k{k}/"
    
    if verbose > 0:
        print(_path_cat)

    CATEGORIES = sorted(os.listdir(_path_cat))
    num_classes = len(CATEGORIES)
    
    if verbose > 0:
        print('classes ', CATEGORIES)
        print('num_classes', num_classes)
        print("\n2-[INFO] Categories and number of classes", num_classes)
    
    return CATEGORIES


def process_train(_config_train, verbose=1):
    """
    Process a single partition of data for training and evaluation.

    Parameters
    ----------
    _config_train : dict
        Configuration dictionary containing parameters for training and evaluation.
    verbose : int, optional
        The verbosity level (default is 1). If verbose > 0, the function prints information.

    Returns
    -------
    model_tl : Model
        The trained model.
    val_loss : float
        The validation loss.
    val_accuracy : float
        The validation accuracy.
    history : History
        The history object returned from the training process.
    start_time : float
        The start time of the training process.
    end_time : float
        The end time of the training process.
    duration : float
        The duration of the training process.
    best_epoch : int
        The epoch with the best validation performance.

    Notes
    -----
    This function performs the following steps:
    1. Hyper model initialization
    2. Load train and validation data
    3. Start training
    4. Evaluate the model
    5. Save the model
    """
    
    if verbose > 0:
        print("\n3-[INFO] Hyper model ")
        
    img_size = _config_train['img_size']
    input_shape = (img_size, img_size, 3)
    input_size = (img_size, img_size)
    
    if verbose > 0:
        print("input_shape ", input_shape)
    
    _config_model = {
        'model': _config_train['model'],
        'id_test': _config_train['id_test'],
        'num_classes': _config_train['num_classes'],
        'last_activation': _config_train['last_activation'],
        'save_dir': _config_train['save_dir'],
        'learning_rate': _config_train['learning_rate'],
        'optimizer': _config_train['optimizer'],
        'input_shape': input_shape,
        'ks': _config_train['ks'],
        'r': _config_train['r']
    }
    
    model_tl = hyper_model(_config_model)

    if verbose > 0:
        print("\n4-[INFO] Load train and validation data")
        
    k = _config_train['k']
    split_valid = float(_config_train['split_valid'])
    
    train, val = get_data.load_data_train(
        _config_train['path_data'],
        K=k,
        BATCH=_config_train['batch_size'],
        INPUT_SIZE=input_size,
        SPLIT_VALID=split_valid
    )

    if verbose > 0:
        print("\n5-[INFO] Start training ")
        
    nm_model = f"{_config_train['id_test']}_{_config_train['model']}"
    
    if verbose > 0:
        print('id_test ', 'Table_Exp', ' nm_model ', nm_model, 'k', k)
    
    train_config = {
        'batch_size': _config_train['batch_size'],
        'epochs': _config_train['epochs'],
        'verbosity': 0
    }

    res_train = models_train.run_train(
        train,
        val,
        model_tl,
        train_config
    )
    
    if verbose > 0:
        print("\n[INFO] Finished training ------------", res_train['duration'])
        print("\n6-[INFO] Evaluating ------------")
        
    val_loss, val_accuracy = model_tl.evaluate(val, batch_size=_config_train['batch_size'], verbose=1)
    
    # Saving the model
    save_dir = _config_train['save_dir']
    
    if save_dir != "":
        save_dir = f"{save_dir}/{nm_model}/"
        
        # Create necessary folders in the specified save directory
        utils.create_folders(save_dir, flag=0)
        
        if verbose > 0:
            print(save_dir)
            
        # utils.create_folders(save_dir, flag=0)
        nome_model = os.path.join(save_dir, f"{nm_model}_bestLoss_{k}.keras")
        model_tl.save(nome_model)
    
    return model_tl, val_loss, val_accuracy, res_train

    
def evaluate_model(model_tl, categories, hist, _config_eval):
    """
    Evaluate the trained model on the validation dataset.

    Parameters
    ----------
    model_tl : keras.Model
        The trained model to evaluate.
    categories : list
        List of categories for the evaluation.
    hist : History
        The history object returned from the training process.
    _config_eval : dict
        Configuration dictionary containing evaluation parameters.

    Returns
    -------
    me : dict
        The evaluation metrics and results.

    Notes
    -----
    This function loads the test data, evaluates the model, generates reports,
    and returns the evaluation metrics and results.
    """
    batch_size = _config_eval['batch_size']
    _path_data = _config_eval['path_data']
    img_size = _config_eval['img_size']
    k = _config_eval['k']
    input_size = (img_size, img_size)
    
    # Load test data
    test = get_data.load_data_test_dir(_path_data, k, batch_size, input_size)
    
    # Configure reports
    reports = {
        'model': _config_eval['model'],
        'batch_size': batch_size,
        'save_dir': _config_eval['save_dir'],
        'fold_var': k,
        'id_test': _config_eval['id_test']
    }
    
    # Generate evaluation reports
    me = reports_gen(test, model_tl, categories, hist, reports)
    
    return me
  
def reports_gen(test_data_generator, model, categories, history, reports_config):
    """
    Generate and save evaluation reports for the trained model.

    Parameters
    ----------
    test_data_generator : DataGenerator
        The data generator for the test dataset.
    model : keras.Model
        The trained model to evaluate.
    categories : list
        List of categories for the evaluation.
    history : History
        The history object returned from the training process.
    reports_config : dict
        Configuration dictionary containing parameters for report generation.

    Returns
    -------
    dict
        A dictionary containing evaluation metrics such as test loss, test accuracy, precision, recall, f-score, and kappa.

    Notes
    -----
    This function evaluates the model, generates various reports including confusion matrix, classification report, 
    training metrics plots, and saves these reports to specified directory. It also returns the evaluation metrics.
    """
    save_dir = reports_config['save_dir']
    k = reports_config['fold_var']
    batch_size = reports_config['batch_size']    
    id_test = reports_config['id_test']
    nm_model = f"{id_test}_{reports_config['model']}"

    # Evaluate the model
    (test_loss, test_accuracy) = model.evaluate(test_data_generator, batch_size=batch_size, verbose=1)
    
    # Predict and generate reports
    #predict_data_generator(test_data_generator, model, categories, batch_size, verbose=2)
    y_true, y_pred, df_correct, df_incorrect = reports.predict_data_generator(
        test_data_generator,
        model, 
        categories,
        batch_size,                                                                                                                                          
    )
    
    # Confusion matrix
    matrix_fig, mat = reports.plot_confusion_matrixV4(y_true, y_pred, categories, normalize=None)
    df_mat = pd.DataFrame(mat, index=categories, columns=categories)
    
    # Boxplot, classification report and training metrics
    boxplot_fig = reports.plot_confidence_boxplot(df_correct)
    class_report = reports.generate_classification_report(y_true, y_pred, categories)
    print(history)
    figTrain = reports.plot_training_metricsV2(history)
    
    # Save files if directory is specified
    if save_dir:
        folder_name=f"{nm_model}_reports/"
        save_dir = os.path.join(save_dir, folder_name)
        print(save_dir)
        utils.create_folders(save_dir, flag=0)
        print("save graph")
        df_mat.to_csv(f'{save_dir}/Test_{id_test}_{nm_model}_mat_conf_k{k}.csv')
        df_correct.to_csv(f'{save_dir}/Test_{id_test}_{nm_model}_df_correct_k{k}.csv')
        df_incorrect.to_csv(f'{save_dir}/Test_{id_test}_{nm_model}_df_incorrect_k{k}.csv')
        
        class_report.to_csv(f'{save_dir}/Test_{id_test}_{nm_model}_Class_reports_k{k}.csv')
        
        
        figTrain.savefig(f'{save_dir}/Test_{id_test}_{nm_model}_TrainLoss_k{k}.jpg')
        matrix_fig.savefig(f'{save_dir}/Test_{id_test}_{nm_model}_mat_conf_k{k}.jpg')
        boxplot_fig.savefig(f'{save_dir}/Test_{id_test}_{nm_model}_boxplot_k{k}.jpg')
    
    # Calculate metrics
    me = reports.calculate_metrics(y_true, y_pred)

    # Return evaluation metrics
    me = {
        'test_loss': test_loss,
        'test_accuracy': test_accuracy,
        'precision': me['precision'],
        'recall': me['recall'],
        'fscore': me['fscore'],
        'kappa': me['kappa'],
    }
    
    return me

def save_results(row, k, val_loss, val_accuracy, me, res_train, path_xlsx):
    """
    Save the results of the evaluation to a file.

    Parameters
    ----------
    row : Series
        Data containing information about the current test execution.
    k : int
        Current partition number.
    val_loss : float
        Validation loss from the model evaluation.
    val_accuracy : float
        Validation accuracy from the model evaluation.
    me : dict
        Metrics generated from the evaluation.
    start_time : str
        Start time of the training process.
    end_time : str
        End time of the training process.
    duration : str
        Duration of the training process.
    best_epoch : int
        The epoch with the best validation performance.
    path_xlsx : str
        The path to the Excel file where the results will be saved.

    Returns
    -------
    None
        This function saves the results but does not return any values.

    Notes
    -----
    This function rounds the validation and test accuracy to a specified precision, prepares the data to be saved,
    and appends it to the specified Excel sheet. Additionally, it resets the GPU memory after saving the results.
    """
    print("\n7-[INFO] Saving evaluation ------------")
    prec = 3
    va = round(val_accuracy, prec)
    ta = round(me['test_accuracy'], prec)
    
    script_name = os.path.basename(__file__)

    data = [
        row['id_test'], row['model'], k, val_loss, va,
        me['test_loss'], ta,
        me['precision'], me['recall'], me['fscore'], me['kappa'],
        res_train['start_time'], res_train['end_time'], res_train['duration'], 
        res_train['best_epoch'], res_train['num_epochs'], script_name
    ]
    save_sheet(path_xlsx, data, 'Table_Exp')

    # Clear memory
    print("\n[INFO] Memory clean ------------")
    maneger_gpu.reset_keras()

def initialize_gpu():
    maneger_gpu.monitor_memory_and_run()

def prepare_for_test():
    """
    Prepare the environment for a test execution by resetting the GPU and waiting for memory to be available.

    Parameters
    ----------
    None

    Returns
    -------
    None
        This function does not return any values.

    Notes
    -----
    This function resets the Keras backend and waits for 60 seconds to ensure that GPU memory is available before starting a test.
    """
    maneger_gpu.monitor_memory_and_run()



def execute_test(params, tests, test_index, start_index, num_partitions, start_k, save_dir, path_xlsx):
    """
    Execute the training and evaluation process for a specific test case.

    Parameters
    ----------
    params : dict
        Dictionary containing configuration parameters.
    tests : DataFrame
        DataFrame containing data for each test execution.
    test_index : int
        The current test index.
    start_index : int
        The starting index for the tests.
    num_partitions : int
        The number of partitions for the k-fold cross-validation.
    start_k : int
        The starting k value for the k-fold.
    save_dir : str
        The directory to save the results.
    path_xlsx : str
        The path to the Excel file where the results will be saved.

    Returns
    -------
    None
        This function executes the training and evaluation process but does not return any values.

    Notes
    -----
    This function retrieves the test configuration, processes each partition of the data,
    trains the model, evaluates it, and saves the results to an Excel file.
    """
    test_id = start_index + test_index
    row = tests.loc[test_id]
    print('\n---------------Test execution')
    print(row)
    print('id_test ', row['id_test'])

    for partition_index in range(num_partitions):
        k = start_k + partition_index
        categories = process_categories(row, k)
        split_valid = int(params['split_valid'])
        print(split_valid)
        config_train = build_train_config(row, params, save_dir, k, categories, split_valid)
        model_tl, val_loss, val_accuracy, res_train = process_train(
            config_train)
        
        reset_environment()
        config_eval = build_eval_config(row, params, save_dir, k)
        history=res_train['history']
        me = evaluate_model(model_tl, categories, history, config_eval)
        del model_tl, history  # Libera as variáveis
        reset_environment()
        
        save_results(row, k, val_loss, val_accuracy, me, res_train, path_xlsx)


def build_train_config(row, params, save_dir, k, categories, split_valid):
    return {
        'model': row['model'],
        'id_test': row['id_test'],
        'path_data': row['path_data'],
        'batch_size': row['batch_size'],
        'dense_size': row['dense_size'],
        'img_size': params['img_size'],
        'num_classes': len(categories),
        'split_valid': params['split_valid'],
        'last_activation': row['last_activation'],
        'save_dir': save_dir,
        'learning_rate': params['learning_rate'],
        'optimizer': row['optimizer'],
        'epochs': params['epochs'],
        'ks': row['ks'],
        'r': row['r'],
        'k': k
    }

def build_eval_config(row, params, save_dir, k):
    return {
        'model': row['model'],
        'id_test': row['id_test'],
        'path_data': row['path_data'],
        'batch_size': row['batch_size'],                
        'save_dir': save_dir,
        'img_size': params['img_size'],
        'k': k        
    }

def reset_environment():
    maneger_gpu.reset_keras()

def run(path_sheet):
    """
    Run the training and evaluation process for the model based on the given configuration.

    Parameters
    ----------
    path_sheet : str
        The full path to the Excel sheet.

    Configuration Parameters in 'params'
    ------------------------------------
    last_test_index : int
        The index of the last test. This allows resuming from the last completed test in case of interruptions.
    k_fold_number : int
        The k-fold number.
    num_k_folds : int
        The number of k-folds for testing.
    num_tests : int
        The number of tests to be performed.

    Returns
    -------
    None
        This function executes the training and evaluation process but does not return any values.

    Notes
    -----
    This function initializes the GPU, reads the configuration and test parameters from the specified sheet,
    and then iteratively executes the training and evaluation process for each test case.
    These controls are essential for resuming the process from the last successful test in case of unexpected interruptions like power failures or system crashes.
    """
    initialize_gpu()
    
    sheet_exists(path_sheet)
    save_dir = os.path.dirname(path_sheet)
    params, tests = read_sheet(path_sheet)

    # Print params to debug the issue
    print("Params:", params)

    num_tests = params["num_tests"]
    last_test_index = params["last_test_index"]
    num_k_folds = params["num_k_folds"]
    k_fold_number = params["k_fold_number"]

    for test_index in range(num_tests):
        prepare_for_test()
        execute_test(params, tests, test_index, last_test_index, num_k_folds, k_fold_number, save_dir, path_sheet)
        reset_environment()

if __name__ == "__main__":
    """
    Main entry point for executing the script. Runs the training and evaluation process
    and provides notification upon completion.

    Parameters
    ----------
    path_sheet : str, optional
        The full path to the Excel sheet. If not provided, a default path is used.

    Returns
    -------
    None
    """
    # Define the default path for the sheet
    default_path_sheet = "./results/phase2/reports_vgg19+cbam/config_FT_VGG19+CBAM_cr_131224.xlsx"

    # Argument parser
    parser = argparse.ArgumentParser(description="Run the training and evaluation process.")
    parser.add_argument("path_sheet", nargs='?', default=default_path_sheet, type=str, help="The full path to the Excel sheet.")
    args = parser.parse_args()
    """_exemple

        python FT_DFT_K10_xlsx.py "/media/jczars/4C22F02A22F01B22/$WinREAgent/Pollen_classification_view/2_fine_tuned/Reports/config_FT_vistas_121124.xlsx"
    """
    

    debug = True
    
    if debug:
        # Run the training and evaluation process in debug mode
        run(args.path_sheet)
        sound_test_finalizado.beep(2)
    else:        
        try:
            # Run the training and evaluation process and send success notification
            run(args.path_sheet)
            sound_test_finalizado.beep(2)
            message = '[INFO] successfully!'
            send_whatsApp_msn.send(message)
            
        except Exception as e:
            # Send error notification if the process fails
            sound_test_finalizado.beep(2)
            message = f'[INFO] with ERROR!!! {str(e)}'
            send_whatsApp_msn.send(message)
