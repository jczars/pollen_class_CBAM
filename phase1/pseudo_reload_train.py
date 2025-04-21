import argparse
import os, sys
import openpyxl
import tensorflow as tf
import pandas as pd
import os

#Variables the environment
# Add the current directory to the PYTHONPATH
sys.path.insert(0, os.getcwd())

# Importing modules and functions
from models import get_data, utils, models_pre, models_train, reports_build
from models import get_calssifica, maneger_gpu, sound_test_finalizado

# Configuring TensorFlow to reduce verbosity
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
tf.get_logger().setLevel('ERROR')  # Limits TensorFlow messages to errors only



"""
Modification to be made: attempt to reduce memory consumption!
"""


def prepare_data(conf, root_path):
    """
    Prepares the environment for the pollen classification experiment.

    Parameters:
        conf (dict): Configuration dictionary containing:
            - id_test (str): Test identifier.
            - model (str): Model name to be used.
            - aug (str): Data augmentation method.
            - base (str): Base folder name of the dataset.
            - path_base (str): Path to the base dataset folder.
        root_path (str): Path to the root directory for results.

    Returns:
        dict: Dictionary with paths and experiment information.
    """
    # Destructuring the configuration dictionary for clarity
    id_test = int(conf['id_test'])
    model = conf['model']
    aug = conf['aug']
    base = conf['base']
    base_path = conf['path_base']

    # Label directory path
    labels_dir = os.path.join(base_path, "labels")
    categories = sorted(os.listdir(labels_dir))

    # Experiment name and path
    experiment_name = f"{id_test}_{model}_{aug}_{base}"
    experiment_path = os.path.join(root_path, experiment_name)
    pseudo_csv_dir = os.path.join(experiment_path, 'pseudo_csv')

    # Creating the necessary folders in a more concise way
    for directory in [root_path, experiment_path, pseudo_csv_dir]:
        utils.create_folders(directory, flag=0)

    print(f"Training save directory: {experiment_path}, ID: {experiment_name}")
    
    # Path to CSV file
    csv_file_path = os.path.join(base_path, f"{base}.csv")
    print('CSV data path:', csv_file_path)
    
    # Creating the labeled dataset
    labeled_data = utils.create_dataSet(labels_dir, csv_file_path, categories)
    
    num_labels = len(labeled_data)
    print('Total labeled data count:', num_labels)
    
    # Splitting data into training, validation and testing
    train_path, val_path, test_path = get_data.splitData(labeled_data, root_path, base)

    return {
        'path_train': train_path,
        'path_val': val_path,
        'path_test': test_path,
        'save_dir_train': experiment_path,
        'pseudo_csv': pseudo_csv_dir,
        'size_of_labels': num_labels,
        'categories': categories
    }


import pandas as pd

def load_data_labels(conf):
    """
    Load training and validation data from CSV files.

    Parameters:
        conf (dict): Configuration dictionary containing:
            - path_train (str): Path to the training data CSV.
            - path_val (str): Path to the validation data CSV.
            - img_size (int): Image size for resizing.
            - aug (str): Data augmentation type.

    Returns:
        tuple: A tuple containing the training data generator, validation data generator, 
               and the original training data DataFrame.

    Raises:
        ValueError: If any required configuration is missing or data loading fails.
    """
    # Check if the required keys are present in the configuration
    required_keys = ['path_train', 'path_val', 'img_size', 'aug']
    missing_keys = [key for key in required_keys if key not in conf]
    if missing_keys:
        raise ValueError(f"Missing required configuration keys: {', '.join(missing_keys)}")

    # Load the training data
    try:
        training_data = pd.read_csv(conf['path_train'])
    except (FileNotFoundError, pd.errors.EmptyDataError) as e:
        raise ValueError(f"Error loading training data: {e}")

    # Load validation data
    try:
        val_data = pd.read_csv(conf['path_val'])
    except (FileNotFoundError, pd.errors.EmptyDataError) as e:
        raise ValueError(f"Error loading validation data: {e}")

    # Get image size
    img_size = conf['img_size']
    if not img_size:
        raise ValueError("img_size is not specified or invalid in the configuration")

    # Set the input size for data generators
    input_size = (img_size, img_size)

    # Load training and validation data with augmentation type
    try:
        train, val = get_data.load_data_train_aug_param(training_data, val_data, conf['aug'], input_size)
                             #load_data_train_aug_param(training_data, val_data, aug, input_size)
    except Exception as e:
        raise ValueError(f"Error during data loading and augmentation: {e}")

    # Return the data generators and training data
    return train, val

def build_train_config(row, res_pre, time_step):
    """Build a configuration dictionary for model training.

    Parameters
    ----------
    row : pandas.Series
        A row from the configuration DataFrame, containing model parameters.
    res_pre : dict
        A dictionary containing the result of the previous step, including labeled data paths.
    iteration_num : int
        The current iteration number in the training loop.

    Returns
    -------
    config : dict
        A dictionary containing the configuration for model training.
    """
    # Extract directories from res_pre
    save_dir_train = res_pre["save_dir_train"]
    test_path = res_pre["path_test"]
    categories = res_pre["categories"]

    # Path where the models will be saved
    save_dir = os.path.join(save_dir_train, "models")

    # Create the training configuration dictionary

    config = {
        "model": row["model"],
        "id_test": row["id_test"],
        "data_path": row["path_base"],
        "test_path": test_path,
        "batch_size": row["batch_size"],
        "img_size": row["img_size"],
        "num_classes": len(categories),
        "split_valid": row["split_valid"],
        "last_activation": row["last_activation"],
        "save_dir": save_dir,
        "learning_rate": row["learning_rate"],
        "optimizer": row["optimizer"],
        "epochs": row["epochs"],
        "freeze": row["freeze"],
        "time_step": time_step,
    }

    return config

def train_model(config, train_data, val_data, time_step):
    """
    Train a model with the given configuration and data. If time_step > 0, it will load a pre-trained model
    and continue training. Otherwise, it will train a new model from scratch.

    Parameters
    ----------
    config : dict
        A dictionary containing the configuration for model training.
    train_data : tuple
        A tuple containing the training data and labels.
    val_data : tuple
        A tuple containing the validation data and labels.
    time_step : int
        The current training step. If time_step > 0, it will load a model from the previous step.

    Returns
    -------
    model_inst : keras.Model
        The trained or reloaded model instance.
    res_train : dict
        A dictionary containing the training history and metrics.
    """
    print('\n[INFO]--> time_step ', time_step)
    
    # Reset model_inst to ensure it starts fresh for each time step
    model_inst = None
    
    # If time_step > 0, try to load the model from a previous step, otherwise train a new model
    if time_step > 0:
        # Build the model path for the previous step
        model_name = f"{config['id_test']}_{config['model']}_bestLoss_{time_step - 1}.keras"
        save_path = os.path.join(config['save_dir'], model_name)
        
        # Load the model from the previous time step
        if os.path.exists(save_path):
            print(f"[INFO]--> Loading model from {save_path}")
            model_inst = tf.keras.models.load_model(save_path)
            
            # Explicitly freeze the layers again
            for layer in model_inst.layers:
                layer.trainable = False  # Freeze all layers
            
            # Optionally unfreeze the last few layers if needed (based on config['freeze'])
            for i, layer in enumerate(model_inst.layers):
                if i >= config['freeze']:
                    layer.trainable = True  # Unfreeze layers after the specified freeze index
            
            print(f"[INFO]--> Model layers frozen up to layer {config['freeze']}")
        else:
            raise ValueError(f"[ERROR]--> Model from time_step {time_step - 1} not found at {save_path}")
    else:
        # Instantiate the model from scratch for time_step == 0
        print("[INFO]--> Training a new model from scratch...")
        model_inst = models_pre.hyper_model_up(config, verbose=1)
    
    # Train the model with the training and validation data
    res_train = models_train.run_train(train_data, val_data, model_inst, config)
    
    # Save the model at the current time step
    model_name = f"{config['id_test']}_{config['model']}_bestLoss_{time_step}.keras"
    save_path = os.path.join(config['save_dir'], model_name)
    
    # Save the trained model
    model_inst.save(save_path)
    print(f"[INFO]--> Model saved at {save_path}")
    
    return model_inst, res_train


def build_reports_config(time_step, config, res_pre, model_inst, res_train, verbose=0):
    """
    Generates evaluation reports for a model based on given configurations, test data, and training results.

    Parameters:
    - time_step (int/float): The time step or timestamp associated with the evaluation.
    - config (dict): A dictionary containing model and report configuration parameters, such as image size and batch size.
    - res_pre (dict): Contains preprocessing results, including the test data path and category information.
    - model_inst: The trained model instance.
    - res_train (dict): Contains training results, including training history.
    - verbose (int, optional): Level of verbosity for printing messages. Default is 0 (no output).

    Returns:
    - report_metrics (Any): Generated metrics from the report generation process.

    Function Workflow:
    1. Conditionally prints log messages indicating the start of report generation based on verbosity level.
    2. Loads test data from a specified path.
    3. Prepares the input size for data processing.
    4. Loads test data for evaluation.
    5. Creates necessary directories for report storage.
    6. Configures and generates reports using provided data and model.
    """
    if verbose > 0:
        print('\nReports Generation')
        print(f'\n[INFO]--> Step 1.4 - Evaluation Time Step: {time_step}')
    
    # Load test data
    test_data = pd.read_csv(res_pre['path_test'])
    print("\n[INFO]--> res_pre['path_test']", res_pre['path_test'])
    print('\n[INFO]--> test_data.head()', test_data.head())

    img_size = config['img_size']
    input_size = (img_size, img_size)
    
    if verbose > 0:
        print(f'\n[INFO]--> Input size: {input_size}')
    
    # Load processed test data
    test = get_data.load_data_test(test_data, input_size)
    categories = res_pre['categories']
    
    # Create report saving directory
    save_dir = os.path.join(res_pre['save_dir_train'], 'reports')
    utils.create_folders(save_dir, 0)
    
    # Configure report generation settings
    reports_config = {
        'save_dir': save_dir,
        'time': time_step,
        'batch_size': config['batch_size'],
        'id_test': config['id_test'],
        'model': config['model']
    }
    
    # Generate reports
    history = res_train['history']
    report_metrics = reports_build.reports_gen(test, model_inst, categories, history, reports_config)
    
    return report_metrics


def classification(config, res_pre, model, _tempo, verbose=0):
    """
    Classifies unlabeled images and generates pseudo-labels.

    Args:
        config (dict): Model configuration and directory paths.
        res_pre (dict): Previous results, including categories and pseudo-label paths.
        model (torch.nn.Module): Trained model used for predictions.
        _tempo (int): Current iteration of the pseudo-labeling process.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).

    Returns:
        pd.DataFrame: DataFrame containing pseudo-label predictions, or None if data is unavailable.
    """
    # Define the path for the unlabeled dataset
    unlabels_path = os.path.join(config['path_base'], 'images_unlabels')
    batch_size = config['batch_size']
    categories = res_pre['categories']

    params = {
        'unlabels': unlabels_path,
        'img_size': config['img_size'],
        'batch_size': batch_size,
        'categories': categories
    }

    # Load unlabeled data or read CSV with previous pseudo-labels
    if _tempo == 0:
        if verbose > 0:
            print(f"[DEBUG] Loading unlabeled images from directory: {unlabels_path}")
        unlabels_generator = get_data.load_unlabels(params)
    else:
        unlabels_csv_path = os.path.join(res_pre['pseudo_csv'], f'unlabelSet_T{_tempo}.csv')
        if verbose > 0:
            print(f"[DEBUG] Loading pseudo-labels from CSV: {unlabels_csv_path}")
        try:
            df_unlabels = pd.read_csv(unlabels_csv_path)
        except FileNotFoundError:
            if verbose > 0:
                print(f"[ERROR] File not found: {unlabels_csv_path}")
            return None

        if len(df_unlabels) > 0:
            if verbose > 0:
                print("[DEBUG] Head of the pseudo-labels DataFrame:")
                print(df_unlabels.head())
            unlabels_generator = get_data.load_data_test(df_unlabels, input_size=(224, 224))
        else:
            if verbose > 0:
                print(f"[WARNING] No data found in CSV {unlabels_csv_path}")
            return None

    # Perform predictions for pseudo-labeling
    if verbose > 0:
        print("[INFO] Performing pseudo-labeling on the unlabeled dataset")
    
    pseudos_df = reports_build.predict_unlabeled_data(
        unlabels_generator, model, batch_size, categories, verbose=verbose
    )
    
    if verbose > 0:
        print(f"[INFO] Total pseudo-labels generated: {len(pseudos_df)}")

    return pseudos_df


def selection(pseudos_df, conf, res_pre, _tempo, verbose=0):
    """
    Performs selection of pseudo-labels for training if unlabeled data is available.

    Steps:
    1. Checks if there is any unlabeled data.
    2. Calls the `selec` function to select pseudo-labels based on a confidence threshold.
    3. Returns a dictionary with paths and sizes of datasets if selection is successful.
    4. Returns None if no selection could be made or if there is no unlabeled data.

    Parameters:
        pseudos_df (DataFrame): Unlabeled data to be processed.
        conf (dict): Configuration dictionary with paths and threshold settings.
        res_pre (dict): Contains the path for saving pseudo-labels.
        _tempo (int): Current time or iteration index.
        training_data (Any): Training data used for comparison or updating with pseudo-labels.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).

    Returns:
        dict or None: Returns a dictionary with new data paths and dataset sizes if successful; 
                      returns None if no selection was made or no unlabeled data.
    """
    if not pseudos_df.empty:
        if verbose > 0:
            print('\n[STEP 2].4 - Selection')

        # Load the training data 
        try:
            training_data = pd.read_csv(res_pre['path_train'])
        except (FileNotFoundError, pd.errors.EmptyDataError) as e:
            raise ValueError(f"Error loading training data: {e}")


        if verbose > 0:
            print(f'Training data provided: {training_data}')

        # Perform pseudo-label selection
        res_sel = get_calssifica.selec(
            conf,
            pseudos_df,
            res_pre['pseudo_csv'], 
            _tempo, 
            training_data,
            conf['limiar']
        )

        if res_sel:
            # Return paths and dataset sizes for further processing
            return {
                '_csv_New_TrainSet': res_sel['_csv_New_TrainSet'],
                'path_test': res_pre['path_test'],
                'save_dir_train': conf.get('path_model', ''),  # Assuming model save path in config
                'pseudo_csv': res_pre['pseudo_csv'],
                'ini': res_sel['ini'],
                'select': res_sel['select'],
                'rest': res_sel['rest'],
                'train': res_sel['train'],
                'new_train': res_sel['new_train']
            }
        else:
            # No valid pseudo-labels selected
            if verbose > 0:
                print("[INFO] No valid pseudo-labels were selected.")
            return None
    else:
        # No unlabeled data to process
        if verbose > 0:
            print("[INFO] No unlabeled data available for processing.")
        return None


def rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, config_index, verbose=0):
    """
    Saves data into an Excel workbook for reporting purposes.

    Parameters:
        time_step (str/int): Current time step or identifier for data logging.
        report_metrics (dict): Metrics from the report generation process.
        res_train (dict): Training result data, including timing and accuracy metrics.
        res_sel (dict): Selection result data, containing training set sizes and other statistics.
        workbook_path (str): Path to the Excel workbook.
        config_index (str/int): Identifier for the configuration used.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).
    """
    if verbose > 0:
        print("\n[INFO] Workbook name:", workbook_path)
    
    try:
        workbook = openpyxl.load_workbook(workbook_path)
        if verbose > 0:
            print("Sheets in workbook:", workbook.sheetnames)
    except FileNotFoundError:
        if verbose > 0:
            print("[ERROR] Workbook not found, creating a new one.")
        workbook = openpyxl.Workbook()

    sheet_name = 'Table'
    
    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        if verbose > 0:
            print(f'Sheet "{sheet_name}" exists.')
        Met_page = workbook[sheet_name]  # Access the existing sheet
    else:
        if verbose > 0:
            print(f'Creating new sheet: "{sheet_name}".')
        Met_page = workbook.create_sheet(sheet_name)  # Create a new sheet
        if verbose > 0:
            print('[INFO] -rel_data- Saving test header.')
        cols_exe = ['Tempo', 'test_loss', 'test_accuracy', 'precision', 'recall', 'fscore', 
                    'kappa', 'str_time', 'end_time', 'delay', 'best_epoch', 
                    'ini', 'select', 'rest', 'train', 'new_train', 'id_test']
        Met_page.append(cols_exe)  # Add header row with column names
    
    # Append data to the sheet
    data = [
        str(time_step),
        report_metrics.get('test_loss', ''),
        report_metrics.get('test_accuracy', ''),
        report_metrics.get('precision', ''),
        report_metrics.get('recall', ''),
        report_metrics.get('fscore', ''),
        report_metrics.get('kappa', ''),
        res_train.get('start_time', ''),
        res_train.get('end_time', ''),
        res_train.get('duration', ''),
        res_train.get('best_epoch', ''),
        res_sel.get('ini', ''),
        res_sel.get('select', ''),
        res_sel.get('rest', ''),
        res_sel.get('train', ''),
        res_sel.get('new_train', ''),
        config_index
    ]
    Met_page.append(data)
    
    # Save the workbook
    workbook.save(workbook_path)
    
    if verbose > 0:
        print("Data saved successfully. Sheets available:", workbook.sheetnames)


def run(workbook_path, start_index, end_index=None, verbose=0):
    """
    Executes the pseudo-labeling and training process based on configurations provided in an Excel sheet.
    
    Args:
        workbook_path (str): Path to the Excel workbook containing configurations.
        start_index (int): Starting index for processing configurations.
        verbose (int, optional): Verbosity level for logging messages. Default is 0 (no output).
    """

    if verbose > 0:
        print("\n[INFO] Workbook name:", workbook_path)
    
    workbook = openpyxl.load_workbook(workbook_path)
    if verbose > 0:
        print("Sheets in workbook:", workbook.sheetnames)

    # Set root path
    root_path = os.path.dirname(workbook_path)
    
    # Load configuration data from 'Sheet'
    config_data = pd.read_excel(workbook_path, sheet_name="Sheet")
    
    # Validate start_index
    if start_index >= len(config_data):
        if verbose > 0:
            print(f"[ERROR] start_index ({start_index}) is out of range.")
        return
    if end_index is None:
        end_index = len(config_data) - start_index

    # Iterate over each row of configuration starting from `start_index`
    for row_idx in range(end_index):
        #utils.reset_keras()  # Reset Keras backend to avoid memory issues, if necessary
        maneger_gpu.monitor_memory_and_run()

        config_index = start_index + row_idx
        config = config_data.loc[config_index]
        if verbose > 0:
            print("Current configuration:", config)

        # Initialize time control and unlabeled data flag
        time_step = 0
        has_unlabeled_data = True
        res_pre = prepare_data(config, root_path)
        
        conf_load = {
            'path_train': res_pre['path_train'],
            'path_val': res_pre['path_val'],
            'img_size': config['img_size'],
            'aug': config['aug']
        }
        if verbose > 0:
            print("Loading data with config:", conf_load)

        train, val = load_data_labels(conf_load)
        del conf_load
        maneger_gpu.log_memory_usage('conf_load')

        while has_unlabeled_data:
            maneger_gpu.monitor_memory_and_run()

            if verbose > 0:
                print('\n[STEP] Training phase')
            conf_train = build_train_config(config, res_pre, time_step)
            model_train, res_train = train_model(conf_train, train, val, time_step)
            del conf_train
            maneger_gpu.log_memory_usage('conf_train')

            report_metrics = build_reports_config(time_step, config, res_pre, model_train, res_train)

            if verbose > 0:
                print('\n[STEP] Classification phase')
            pseudos_df = classification(config, res_pre, model_train, time_step)  
            del model_train 
            maneger_gpu.log_memory_usage('model_train')          
            
            if verbose > 0:
                print('\n[STEP] Selection phase')
            res_sel = selection(pseudos_df, config, res_pre, time_step)
            del pseudos_df
            maneger_gpu.log_memory_usage('training_data')   

            if res_sel is None:
                if verbose > 0:
                    print("[INFO] No more unlabeled data to process.")
                break

            rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, config_index)
            del report_metrics, res_train
            maneger_gpu.log_memory_usage('report_metrics')  
            
            # Reload data if necessary
            if time_step > 0:
                train, val = get_data.reload_data_train(config, res_sel['_csv_New_TrainSet'])
            del res_sel
            maneger_gpu.log_memory_usage('res_sel')

            time_step += 1


if __name__ == "__main__":
    # Configuração do argparse para lidar com argumentos de linha de comando
    default_path='0_pseudo_labels/Reports/config_pseudo_label_pre.xlsx'
    parser = argparse.ArgumentParser(description="Run the pollen classification process.")
    parser.add_argument(
        '--path', 
        type=str, 
        default=default_path, 
        help="Path to the workbook. If not provided, the default path will be used."
    )
    parser.add_argument(
        '--start_index', 
        type=int, 
        default=0, 
        help="Starting index for the run. Default is 0."
    )
    parser.add_argument(
        '--end_index', 
        type=int, 
        default=None, 
        help="Ending index for the run. If not provided, the function will use the default (None)."
    )

    args = parser.parse_args()
    
    # Check if the given path exists
    if not os.path.exists(args.path):
        print(f"Warning: The provided workbook path '{args.path}' does not exist.")
        print("Using default workbook path.")
        args.path = default_path

    # Call the 'run' function with the arguments
    run(args.path, args.start_index, args.end_index)
    sound_test_finalizado.beep(2)

