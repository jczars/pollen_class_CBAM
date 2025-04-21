
# Origin 0_pseudo_labels/pseudo_rec_train_load_b2.ipynb

# Importing necessary libraries
import argparse
import pandas as pd
import openpyxl
import tensorflow as tf
import os, sys
#from memory_profiler import profile

#Variables the environment
# Add the current directory to the PYTHONPATH
sys.path.insert(0, os.getcwd())
# Importing modules and functions
from models import get_data, models_train, get_calssifica, sound_test_finalizado, utils, reports_build

# Configuring TensorFlow to reduce verbosity
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
tf.get_logger().setLevel('ERROR')  # Limits TensorFlow messages to errors only

# # Functions
#path_reports = './results/phase1/reports_cr/'

# ## rec
#@profile
def rec_id(workbook_path, id_test):
    # Load configuration data from 'Sheet'
    """
    Retrieves configuration and counts occurrences of a given test ID from an Excel workbook.

    Parameters:
    workbook_path (str): Path to the Excel workbook containing configuration and test data.
    id_test (int): Identifier for the test configuration and data to be retrieved.

    Returns:
    tuple: 
        - config (pandas.Series): Configuration data for the specified test ID.
        - tempo_px (int): Number of occurrences of the specified test ID in the 'Table' sheet.
    """
    config_data = pd.read_excel(workbook_path, sheet_name="Sheet")
    config = config_data.loc[id_test]
    rec_csv = pd.read_excel(workbook_path, sheet_name="Table")
    fil=rec_csv[rec_csv['id_test'] == id_test]
    tempo_px=len(fil)

    return config, tempo_px


# ## train_model


#@profile
def train_model(config, train_data, val_data, time_step):
    """
    Train a model with the given configuration and data. If time_step > 0, it will load a pre-trained model
    and continue training. Otherwise, it will train a new model from scratch.

    Parameters
    ----------
    config : dict
        A dictionary containing the configuration for model training.
    train_data : tuple
        A tuple containing the training data and labels.
    val_data : tuple
        A tuple containing the validation data and labels.
    time_step : int
        The current training step. If time_step > 0, it will load a model from the previous step.

    Returns
    -------
    model_inst : keras.Model
        The trained or reloaded model instance.
    res_train : dict
        A dictionary containing the training history and metrics.
    """
    print('\n[INFO]--> time_step ', time_step)
    
    # Reset model_inst to ensure it starts fresh for each time step
    model_inst = None
    
    # If time_step > 0, try to load the model from a previous step, otherwise train a new model
    if time_step > 0:
        # Build the model path for the previous step
        model_name = f"{config['id_test']}_{config['model']}_bestLoss_{time_step - 1}.keras"
        save_path = os.path.join(config['save_dir'], model_name)
        
        # Load the model from the previous time step
        if os.path.exists(save_path):
            print(f"[INFO]--> Loading model from {save_path}")
            model_inst = tf.keras.models.load_model(save_path)
            
            # Explicitly freeze the layers again
            for layer in model_inst.layers:
                layer.trainable = False  # Freeze all layers
            
            # Optionally unfreeze the last few layers if needed (based on config['freeze'])
            for i, layer in enumerate(model_inst.layers):
                if i >= config['freeze']:
                    layer.trainable = True  # Unfreeze layers after the specified freeze index
            
            print(f"[INFO]--> Model layers frozen up to layer {config['freeze']}")
        else:
            raise ValueError(f"[ERROR]--> Model from time_step {time_step - 1} not found at {save_path}")
    
    # else:
    #     # Instantiate the model from scratch for time_step == 0
    #     print("[INFO]--> Training a new model from scratch...")
    #     model_inst = models_pre.hyper_model_up(config, verbose=1)
    
    # Train the model with the training and validation data
    res_train = models_train.run_train(train_data, val_data, model_inst, config)
    
    # Save the model at the current time step
    model_name = f"{config['id_test']}_{config['model']}_bestLoss_{time_step}.keras"
    save_path = os.path.join(config['save_dir'], model_name)
    
    # Save the trained model
    model_inst.save(save_path)
    print(f"[INFO]--> Model saved at {save_path}")
    
    return model_inst, res_train


# ## build_reports_config


#@profile
def build_reports_config(time_step, config, res_pre, model_inst, res_train, verbose=0):
    """
    Generates evaluation reports for a model based on given configurations, test data, and training results.

    Parameters:
    - time_step (int/float): The time step or timestamp associated with the evaluation.
    - config (dict): A dictionary containing model and report configuration parameters, such as image size and batch size.
    - res_pre (dict): Contains preprocessing results, including the test data path and category information.
    - model_inst: The trained model instance.
    - res_train (dict): Contains training results, including training history.
    - verbose (int, optional): Level of verbosity for printing messages. Default is 0 (no output).

    Returns:
    - report_metrics (Any): Generated metrics from the report generation process.

    Function Workflow:
    1. Conditionally prints log messages indicating the start of report generation based on verbosity level.
    2. Loads test data from a specified path.
    3. Prepares the input size for data processing.
    4. Loads test data for evaluation.
    5. Creates necessary directories for report storage.
    6. Configures and generates reports using provided data and model.
    """
    if verbose > 0:
        print('\nReports Generation')
        print(f'\n[INFO]--> Step 1.4 - Evaluation Time Step: {time_step}')
    
    # Load test data
    test_data = pd.read_csv(res_pre['path_test'])
    print("\n[INFO]--> res_pre['path_test']", res_pre['path_test'])
    print('\n[INFO]--> test_data.head()', test_data.head())

    img_size = config['img_size']
    input_size = (img_size, img_size)
    
    if verbose > 0:
        print(f'\n[INFO]--> Input size: {input_size}')
    
    # Load processed test data
    test = get_data.load_data_test(test_data, input_size)
    categories = res_pre['categories']
    
    # Create report saving directory
    save_dir = os.path.join(res_pre['save_dir_train'], 'reports')
    utils.create_folders(save_dir, 0)
    
    # Configure report generation settings
    reports_config = {
        'save_dir': save_dir,
        'time': time_step,
        'batch_size': config['batch_size'],
        'id_test': config['id_test'],
        'model': config['model']
    }
    
    # Generate reports
    history = res_train['history']
    report_metrics = reports_build.reports_gen(test, model_inst, categories, history, reports_config)
    
    return report_metrics


# ## classification


#@profile
def classification(config, res_pre, model, _tempo, verbose=0):
    """
    Classifies unlabeled images and generates pseudo-labels.

    Args:
        config (dict): Model configuration and directory paths.
        res_pre (dict): Previous results, including categories and pseudo-label paths.
        model (torch.nn.Module): Trained model used for predictions.
        _tempo (int): Current iteration of the pseudo-labeling process.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).

    Returns:
        pd.DataFrame: DataFrame containing pseudo-label predictions, or None if data is unavailable.
    """
    # Define the path for the unlabeled dataset
    unlabels_path = os.path.join(config['path_base'], 'images_unlabels')
    batch_size = config['batch_size']
    categories = res_pre['categories']

    params = {
        'unlabels': unlabels_path,
        'img_size': config['img_size'],
        'batch_size': batch_size,
        'categories': categories
    }

    # Load unlabeled data or read CSV with previous pseudo-labels
    if _tempo == 0:
        if verbose > 0:
            print(f"[DEBUG] Loading unlabeled images from directory: {unlabels_path}")
        unlabels_generator = get_data.load_unlabels(params)
    else:
        unlabels_csv_path = os.path.join(res_pre['pseudo_csv'], f'unlabelSet_T{_tempo}.csv')
        if verbose > 0:
            print(f"[DEBUG] Loading pseudo-labels from CSV: {unlabels_csv_path}")
        try:
            df_unlabels = pd.read_csv(unlabels_csv_path)
        except FileNotFoundError:
            if verbose > 0:
                print(f"[ERROR] File not found: {unlabels_csv_path}")
            return None

        if len(df_unlabels) > 0:
            if verbose > 0:
                print("[DEBUG] Head of the pseudo-labels DataFrame:")
                print(df_unlabels.head())
            unlabels_generator = get_data.load_data_test(df_unlabels, input_size=(224, 224))
        else:
            if verbose > 0:
                print(f"[WARNING] No data found in CSV {unlabels_csv_path}")
            return None

    # Perform predictions for pseudo-labeling
    if verbose > 0:
        print("[INFO] Performing pseudo-labeling on the unlabeled dataset")
    
    pseudos_df = reports_build.predict_unlabeled_data(
        unlabels_generator, model, batch_size, categories, verbose=verbose
    )
    
    if verbose > 0:
        print(f"[INFO] Total pseudo-labels generated: {len(pseudos_df)}")

    return pseudos_df


# ## selection


#@profile
def selection(pseudos_df, conf, res_pre, _tempo, verbose=0):
    """
    Performs selection of pseudo-labels for training if unlabeled data is available.

    Steps:
    1. Checks if there is any unlabeled data.
    2. Calls the `selec` function to select pseudo-labels based on a confidence threshold.
    3. Returns a dictionary with paths and sizes of datasets if selection is successful.
    4. Returns None if no selection could be made or if there is no unlabeled data.

    Parameters:
        pseudos_df (DataFrame): Unlabeled data to be processed.
        conf (dict): Configuration dictionary with paths and threshold settings.
        res_pre (dict): Contains the path for saving pseudo-labels.
        _tempo (int): Current time or iteration index.
        training_data (Any): Training data used for comparison or updating with pseudo-labels.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).

    Returns:
        dict or None: Returns a dictionary with new data paths and dataset sizes if successful; 
                      returns None if no selection was made or no unlabeled data.
    """
    if not pseudos_df.empty:
        if verbose > 0:
            print('\n[STEP 2].4 - Selection')

        # Load the training data 
        try:
            training_data = pd.read_csv(res_pre['path_train'])
        except (FileNotFoundError, pd.errors.EmptyDataError) as e:
            raise ValueError(f"Error loading training data: {e}")


        if verbose > 0:
            print(f'Training data provided: {training_data}')

        # Perform pseudo-label selection
        res_sel = get_calssifica.selec(
            conf,
            pseudos_df,
            res_pre['pseudo_csv'], 
            _tempo, 
            training_data,
            conf['limiar']
        )

        if res_sel:
            # Return paths and dataset sizes for further processing
            return {
                '_csv_New_TrainSet': res_sel['_csv_New_TrainSet'],
                'path_test': res_pre['path_test'],
                'save_dir_train': conf.get('path_model', ''),  # Assuming model save path in config
                'pseudo_csv': res_pre['pseudo_csv'],
                'ini': res_sel['ini'],
                'select': res_sel['select'],
                'rest': res_sel['rest'],
                'train': res_sel['train'],
                'new_train': res_sel['new_train']
            }
        else:
            # No valid pseudo-labels selected
            if verbose > 0:
                print("[INFO] No valid pseudo-labels were selected.")
            return None
    else:
        # No unlabeled data to process
        if verbose > 0:
            print("[INFO] No unlabeled data available for processing.")
        return None


#@profile
def rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, config_index, verbose=0):
    """
    Saves data into an Excel workbook for reporting purposes.

    Parameters:
        time_step (str/int): Current time step or identifier for data logging.
        report_metrics (dict): Metrics from the report generation process.
        res_train (dict): Training result data, including timing and accuracy metrics.
        res_sel (dict): Selection result data, containing training set sizes and other statistics.
        workbook_path (str): Path to the Excel workbook.
        config_index (str/int): Identifier for the configuration used.
        verbose (int, optional): Verbosity level for printing messages. Default is 0 (no output).
    """
    if verbose > 0:
        print("\n[INFO] Workbook name:", workbook_path)
    
    try:
        workbook = openpyxl.load_workbook(workbook_path)
        if verbose > 0:
            print("Sheets in workbook:", workbook.sheetnames)
    except FileNotFoundError:
        if verbose > 0:
            print("[ERROR] Workbook not found, creating a new one.")
        workbook = openpyxl.Workbook()

    sheet_name = 'Table'
    
    # Check if the sheet already exists
    if sheet_name in workbook.sheetnames:
        if verbose > 0:
            print(f'Sheet "{sheet_name}" exists.')
        Met_page = workbook[sheet_name]  # Access the existing sheet
    else:
        if verbose > 0:
            print(f'Creating new sheet: "{sheet_name}".')
        Met_page = workbook.create_sheet(sheet_name)  # Create a new sheet
        if verbose > 0:
            print('[INFO] -rel_data- Saving test header.')
        cols_exe = ['Tempo', 'test_loss', 'test_accuracy', 'precision', 'recall', 'fscore', 
                    'kappa', 'str_time', 'end_time', 'delay', 'best_epoch', 
                    'ini', 'select', 'rest', 'train', 'new_train', 'id_test']
        Met_page.append(cols_exe)  # Add header row with column names
    
    # Append data to the sheet
    data = [
        str(time_step),
        report_metrics.get('test_loss', ''),
        report_metrics.get('test_accuracy', ''),
        report_metrics.get('precision', ''),
        report_metrics.get('recall', ''),
        report_metrics.get('fscore', ''),
        report_metrics.get('kappa', ''),
        res_train.get('start_time', ''),
        res_train.get('end_time', ''),
        res_train.get('duration', ''),
        res_train.get('best_epoch', ''),
        res_sel.get('ini', ''),
        res_sel.get('select', ''),
        res_sel.get('rest', ''),
        res_sel.get('train', ''),
        res_sel.get('new_train', ''),
        config_index
    ]
    Met_page.append(data)
    
    # Save the workbook
    workbook.save(workbook_path)
    
    if verbose > 0:
        print("Data saved successfully. Sheets available:", workbook.sheetnames)


# # Main


def run(workbook_path: str, id_test: int, verbose=1):
    """
    Main function to run pseudo-labelling process.

    Parameters
    ----------
    workbook_path : str
        Path to the workbook where data is stored.
    id_test : int
        Identifier for the current test.
    verbose : int (optional)
        Verbosity level. Defaults to 1.

    Notes
    -----
    The following steps are performed in order:
        1. Recovery phase: recover the configuration and time step.
        2. Train phase: train a model using the current dataset.
        3. Classification phase: classify the pseudo-labels using the trained model.
        4. Selection phase: select the most confident pseudo-labels.
        5. Save reports: save the results of the current iteration.
        6. Repeat the process until no more unlabeled data is available.

    """
    
    has_unlabeled_data = True
    print('\n[STEP] Recovery phase')
    #%memit  
    config, time_step = rec_id(workbook_path, id_test)
    print("\nconfig", config)
    print("\ntempo_px", time_step) 

    while has_unlabeled_data: 
        print(f'*'*30)               
        print('\n[STEP] Train phase')

        confi_load={
            'aug': config['aug'],
            'img_size': config['img_size'],
        }

        # recurarar csv_NewtainSet14
        _pseudo_csv=f'{path_reports}{id_test}_{config["model"]}_{config["aug"]}_{config["base"]}/pseudo_csv/'
        _csv_New_TrainSet = os.path.join(_pseudo_csv, f'trainSet_T{time_step}.csv')
        print(f"_csv_New_TrainSet {_csv_New_TrainSet}")

        save_dir = f'{path_reports}{id_test}_{config["model"]}_{config["aug"]}_{config["base"]}/models/'
        
        config_train={
        'id_test': id_test,
        'model': config['model'],
        'save_dir': save_dir,
        'freeze': config['freeze'],
        'batch_size': config['batch_size'],
        'epochs': config['epochs'],
        }

        train, val = get_data.reload_data_train(confi_load, _csv_New_TrainSet)

        model_train, res_train = train_model(config_train, train, val, time_step)

        # Label directory path
        labels_dir = os.path.join(config['path_base'], "labels")
        categories = sorted(os.listdir(labels_dir))

        class_config={'path_base': config['path_base'],
                    'batch_size': config['batch_size'],
                    'img_size': config['img_size'],
                    'categories': categories,
                    'pseudo_csv': _pseudo_csv
        }
        path_test= f"{path_reports}{config['base']}_testSet.csv"
        train_path= f"{path_reports}{config['base']}_trainSet.csv"
        save_dir_train= f'{path_reports}{id_test}_{config["model"]}_{config["aug"]}_{config["base"]}'
        num_labels = len(categories)
        print(path_test)
        res_pre = {
            'path_train': train_path,
            'path_test': path_test,
            'save_dir_train': save_dir_train,
            'pseudo_csv': _pseudo_csv,
            'size_of_labels': num_labels,
            'categories': categories
        }

        report_metrics = build_reports_config(time_step, config, res_pre, model_train, res_train)
        
        print('\n[STEP] Classification phase')
        pseudos_df = classification(class_config, res_pre, model_train, time_step)  

        print('\n[STEP] Selection phase')
        res_sel = selection(pseudos_df, config, res_pre, time_step)

        if res_sel is None:
            if verbose > 0:
                print("[INFO] No more unlabeled data to process.")
            break
        print('\n[STEP] save reports')
        rel_data(time_step, report_metrics, res_train, res_sel, workbook_path, id_test)
        
        time_step += 1
        print(f'*'*30)   
        print('\n[STEP] Next step', time_step)


if __name__ == "__main__":    
    # Configuração do argparse para lidar com argumentos de linha de comando
    default_path='./results/phase1/reports_cr/config_pseudo_label_pre_cr.xlsx'
    parser = argparse.ArgumentParser(description="Run the pollen classification process.")
    parser.add_argument(
        '--path', 
        type=str, 
        default=default_path, 
        help="Path to the workbook. If not provided, the default path will be used."
    )
    parser.add_argument(
        '--start_index', 
        type=int, 
        default=5, 
        help="Starting index for the run. Default is 5."
    )

    args = parser.parse_args()
    
    # Check if the given path exists
    if not os.path.exists(args.path):
        print(f"Warning: The provided workbook path '{args.path}' does not exist.")
        print("Using default workbook path.")
        args.path = default_path

    # Determine the base directory from the provided path
    base_dir = os.path.dirname(args.path)
    path_reports = os.path.join(base_dir, '')  # Construir dinamicamente o path_reports
    print(f"[INFO] path_reports set to: {path_reports}")

    # Call the 'run' function with the arguments
    run(args.path, args.start_index)
    sound_test_finalizado.beep(2)




